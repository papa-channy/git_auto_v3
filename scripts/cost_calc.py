import requests, json
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from scripts.dataframe import load_df, save_df, PROMPT_PATH

def get_usd_exchange_rate(log) -> float:
    ex_path = Path("cost/ex_rate.txt")
    fallback = 1400.0

    try:
        if ex_path.exists():
            last_modified = datetime.fromtimestamp(ex_path.stat().st_mtime)
            if datetime.now() - last_modified < timedelta(hours=24):
                content = ex_path.read_text().strip()
                if content:
                    return float(content)

        log("🌐 환율 정보 새로 요청 중...")
        url = "https://finance.naver.com/marketindex/"
        html = requests.get(url, timeout=5).text
        soup = BeautifulSoup(html, "html.parser")
        rate_text = soup.select_one("div.head_info > span.value").text
        rate = float(rate_text.replace(",", ""))

        ex_path.parent.mkdir(parents=True, exist_ok=True)
        ex_path.write_text(str(rate), encoding="utf-8")
        return rate

    except Exception as e:
        log(f"⚠️ 환율 정보 가져오기 실패: {e} → fallback {fallback}")
        return fallback


def append_to_excel(df: pd.DataFrame, timestamp: str, filepath: str, log):
    dt = datetime.strptime(timestamp, "%Y%m%d_%H%M")
    df.insert(0, "datetime", dt)

    file = Path(filepath)
    if file.exists():
        wb = load_workbook(file)
        ws = wb.active
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
    else:
        wb = Workbook()
        ws = wb.active
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    wb.save(file)
    log(f"✅ 엑셀 누적 저장 완료 → {filepath}")


def calculate_llm_costs_from_df(timestamp: str, log):
    df = load_df(PROMPT_PATH)
    rate_config = json.loads(Path("config/cost.json").read_text(encoding="utf-8"))
    krw_rate = get_usd_exchange_rate(log)

    usd_list, krw_list = [], []
    for _, row in df.iterrows():
        model = row["사용 MODEL NAME"]
        in_token = row["token값"] if row["IN/OUT"] == "입력" else 0
        out_token = row["token값"] if row["IN/OUT"] == "출력" else 0

        in_rate = rate_config[model]["input"]
        out_rate = rate_config[model]["output"]

        usd = (in_token * in_rate + out_token * out_rate)
        krw = usd * krw_rate
        usd_list.append(round(usd, 6))
        krw_list.append(round(krw))

    df["비용($)"] = usd_list
    df["비용(krw)"] = krw_list

    save_df(df, PROMPT_PATH)
    append_to_excel(df[["datetime", "IN/OUT", "VAR NAME", "사용 MODEL NAME", "token값", "비용($)", "비용(krw)"]], timestamp, "cost/git_cost.xlsx", log)
    return df

